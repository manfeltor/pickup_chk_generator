import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from datetime import datetime
from openpyxl.drawing.image import Image as XLImage
import os

# Function to get path to bundled resource
def resource_path(relative_path):
    try:
        base_path = os.sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

def excels_generator():

    ruteo_df, merged_df = main_pick_generator()
    ruteo_df.to_excel('ruteo.xlsx', index=False)
    merged_df.to_excel('merged.xlsx', index=False)
    remitos_dfs = separate_df_by_sucursal(merged_df)
    remitos_book = generate_excel_with_sheets(remitos_dfs)
    remitos_book.save('remitos.xlsx')
    # add_logo_excel_remitos('remitos.xlsx')

    return None


def main_pick_generator():

    try:
        path_raw_oms = select_file("Selecciona el archivo de OMS")
        df_raw = pd.read_excel(path_raw_oms)
        clean_df = clean_raw_df(df_raw)
        ruteo_df = ruteo_generator(clean_df)
        path_bd_suc = select_file("Selecciona el archivo de la base de sucursales")
        sucursales_df = pd.read_excel(path_bd_suc)
        merged_df = merged_df_generator(clean_df, sucursales_df)

        return ruteo_df, merged_df
    
    except Exception as e:
        messagebox.showerror("Error", f"Error while processing file: {str(e)}")
        return None


def ruteo_generator(clean_df: pd.DataFrame):

    ruteo_df = clean_df[['pedido', 'tiendaEntrega']]
    ruteo_df.sort_values(by='tiendaEntrega', inplace=True)
    ruteo_df.reset_index(drop=True, inplace=True)
    ruteo_df['register_enum'] = ruteo_df.groupby('tiendaEntrega').cumcount() + 1

    return ruteo_df


def select_file(title):

    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(title=title, filetypes=[("Excel files", "*.xlsx")])
    root.destroy()

    if not file_path:
        messagebox.showerror("Error", "No file selected.")
        return None

    return file_path


def clean_raw_df(raw_df: pd.DataFrame):

    try:
        filtered_df = raw_df[['pedido', 'nombre', 'tiendaEntrega']].drop_duplicates(subset=['pedido'])
        return filtered_df
    
    except Exception as e:
        messagebox.showerror("Error", f"Error while cleaning DataFrame: {str(e)}")
        return None
    

def merged_df_generator(clean_df: pd.DataFrame, sucursales_df: pd.DataFrame):
    try:
        filtered_df = clean_df[["pedido", "tiendaEntrega", "nombre"]]
        filtered_df_len = filtered_df.shape[0]
        inner_join = pd.merge(filtered_df, sucursales_df, on='tiendaEntrega', how='inner')
        inner_join_len = inner_join.shape[0] 

        if filtered_df_len == inner_join_len:
            return inner_join
        else:
            messagebox.showerror("Error", f"algo no coincide en la base CA con el archivo OMS")
            return None

    except Exception as e:
        messagebox.showerror("Error", f"Error while merging OMS and CA dbs: {str(e)}")
        return None
    

def separate_df_by_sucursal(merged_df):

    sucursales_dfs = {}
    sucs_codes = list(merged_df['tiendaEntrega'].unique())

    for code in sucs_codes:

        suc_df = merged_df[merged_df['tiendaEntrega'] == code]
        sucursales_dfs[code] = suc_df
    
    return sucursales_dfs


def generate_excel_with_sheets(sucursales_dfs):

    template_file = resource_path('remito_template.xlsx')
    wb = load_workbook(template_file)
    
    for sheet_name, df in sucursales_dfs.items():

        logo = XLImage(resource_path('logo.png'))

        df_len = len(df)
        suc_name = str(df['LOCAL'].unique())
        current_date_str = datetime.now().strftime(r"%Y-%m-%d")
        ws = wb.copy_worksheet(wb['TemplateSheet'])
        ws.title = str(sheet_name)
        ws['A2'] = suc_name[2:-2]
        ws['B4'] = df_len
        ws['B5'] = current_date_str
        ws.add_image(logo, 'C1')
        ws.oddHeader.center.text = " PICKUP CHK - CQ"
        ws.oddHeader.center.size = 15
        ws.oddHeader.center.font = "Tahoma,Bold"

        data_cell_counter = 8
        for index, row in df.iterrows():

            ws.cell(row=(data_cell_counter), column=1).value = row['pedido']
            ws.cell(row=(data_cell_counter), column=2).value = row['nombre']
            ws.cell(row=(data_cell_counter), column=3).value = row['DIRECCION']
            data_cell_counter += 1

    return wb


excels_generator()